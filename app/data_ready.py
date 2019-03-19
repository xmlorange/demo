# -*- coding: utf-8 -*-
"""
    author: Q.Y.
"""
import datetime
from io import BytesIO

import pyexcel as pe
import openpyxl as op
from openpyxl.styles import Alignment, Font

from app.common import NEW_STORE_MAP, date_param, StaticParam
from app.mysql_util import ex


def get_summary_of_absences():
    title = [["学员", "分馆", "上课日期", "任课老师", "电话"]]
    data = []

    key = None
    row_num = 0
    for row in ex.bs("""select se.name,ls.clazz_id,date_format(ls.lesson_start_time,'%Y-%m-%d'),t.name,
                            s.id,s.name,s.mobile,sl.is_signed,sl.id
                            from by_lesson_schedule ls
                        left join by_student_lesson sl on sl.lesson_schedule_id = ls.id
                        left join by_contract_course_allocation bcca on sl.contract_course_allocation_id=bcca.id
                        left join by_contract c on bcca.contract_id = c.id
                        left join by_student s on sl.student_id=s.id
                        left join staff_prod.by_staff t on ls.teacher_id = t.id
                        left join staff_prod.by_store se on ls.store_id=se.id
                        where ls.is_deleted=false and sl.is_deleted=false and c.is_deleted=false 
                            and bcca.is_deleted=false and ls.is_signed = true # and sl.is_signed=false
                            and ls.lesson_start_time < curdate() 
                            and ls.lesson_start_time >= date_add(curdate(),interval -13 day)
                        order by ls.clazz_id,s.id,sl.id desc"""):
        array = list(row)
        if key is None:
            key = (array[1], array[4])
            row_num += 1
            array.append(row_num)
        elif key == (row[1], row[4]):
            row_num += 1
            array.append(row_num)
        else:
            row_num = 1
            key = (row[1], row[4])
            array.append(row_num)

        data.append(array)

    absences_data = {}
    for line in data:
        if line[-1] > 2 or line[-3] == b'\x01':
            continue
        key = (line[1], datetime.datetime.strptime(line[2], '%Y-%m-%d').date(), line[4])
        absences_data[key] = [line[5], line[0], line[2], line[3], line[6]]
    yesterday = (datetime.datetime.now() - datetime.timedelta(days=1)).date()
    for key, value in absences_data.items():
        if key[1] == yesterday:
            _key = (key[0], key[1] - datetime.timedelta(days=7), key[2])
            if absences_data.get(_key, None):
                title.append(value)
    return title


def division(x, y):
    if y == 0:
        return 0
    return round(x / y, 4)


class SignedTable:

    def __init__(self, store_id, store_name, all_signed_mobile, new_signed_mobile):
        self.new_signed_mobile = new_signed_mobile
        self.all_signed_mobile = all_signed_mobile
        self.store_id = store_id
        self.store_name = store_name

        self.ignore_mobiles = self.get_ignore_mobile()

    def run(self):
        old_need_sign_count, old_signed_count, old_new_sign_count = self.get_old_contract()
        new_contract_count, new_signed_count, new_signed_days = self.get_new_contract()

        new_contract_arrange_count, un_arranged_count = self.get_class_new_contract()

        this_week_need_arrange, this_week_arranged = self.get_need_arrange()

        return {
            "学员端情况": {
                "老学员": {
                    "应绑定": old_need_sign_count,
                    "实绑定": old_signed_count,
                    "本周老学员绑定数": old_new_sign_count,
                    "绑定率": division(old_signed_count, old_need_sign_count)
                },
                "新签情况": {
                    "新签合同数": new_contract_count,
                    "学员端绑定数": new_signed_count,
                    "绑定率": division(new_signed_count, new_contract_count),
                    "平均绑定天数": division(new_signed_days, new_signed_count)
                }},

            "排课情况": {
                "本周新签已排课": new_contract_arrange_count,
                "新签总计未排课": un_arranged_count,
                "本周应该排课": this_week_need_arrange,
                "本周实际排课": this_week_arranged
            }

        }

    def get_old_contract(self):
        need_sign_count = 0
        signed_count = 0
        new_sign_count = 0

        for _, contract_no, contract_mobile, student_mobile, student_name in ex.bs("""
            select c.id,c.contract_no,c.mobile,group_concat(distinct s.mobile),group_concat(distinct s.name)
                    from by_contract c 
                        left join by_contract_student cs on c.id =cs.contract_id
                        left join by_student s on s.id = cs.student_id
                    where c.is_deleted=false and cs.is_deleted = false and c.create_time <='{date}' 
                        and c.store_id = {store_id} and c.lesson_cnt > c.lesson_finished_count
                    group by c.id""".format(store_id=self.store_id, date=StaticParam.TIME_POINT)):

            flag = False
            if contract_mobile in self.ignore_mobiles or any(
                    [i in self.ignore_mobiles for i in student_mobile.split(",")]):
                # print(contract_mobile)
                continue
            need_sign_count += 1
            if contract_mobile in self.all_signed_mobile or any(
                    [i in self.all_signed_mobile for i in student_mobile.split(",")]):
                signed_count += 1
                flag = True
            if contract_mobile in self.new_signed_mobile or any(
                    [i in self.new_signed_mobile for i in student_mobile.split(",")]):
                new_sign_count += 1
                flag = True
        return need_sign_count, signed_count, new_sign_count

    def get_new_contract(self):

        contract_count = 0
        signed_count = 0
        signed_days = 0

        for create_time, contract_mobile, student_mobile in ex.bs("""
                select c.create_time,c.mobile,group_concat(DISTINCT bs.mobile) from by_contract c
                  left join by_contract_student bcs on c.id = bcs.contract_id
                  left join by_student bs on bcs.student_id = bs.id
                where c.is_deleted = false and c.create_time >= '{start}' 
                  and c.create_time < '{end}'
                  and bcs.is_deleted=false
                  and c.renew_from_contract_id is null
                  and c.store_id = {store_id}
                group by c.id""".format(store_id=self.store_id,
                                        start=date_param.start,
                                        end=date_param.end)):

            contract_count += 1
            if contract_mobile in self.all_signed_mobile or any(
                    [i in self.all_signed_mobile for i in student_mobile.split(",")]):
                signed_time = self.get_signed_time(contract_mobile, student_mobile.split(",")).date()

                signed_count += 1
                _days = (signed_time - create_time.date()).days

                signed_days += (int(_days) + 1)

        return contract_count, signed_count, signed_days

    def get_class_new_contract(self):

        arranged_count = 0
        un_arranged_count = 0
        last_week = datetime.datetime.combine((datetime.datetime.now() - datetime.timedelta(days=8)).date(),
                                              datetime.time.min)

        for _, lesson_arranged_count, create_time in ex.bs("""
                            select c.id,c.lesson_arraged_count,c.create_time
                                from by_contract c
                            where c.create_time >='{date}' and c.is_deleted=FALSE 
                                and c.renew_from_contract_id is null and c.store_id={store_id}
                            """.format(store_id=self.store_id, date=StaticParam.TIME_POINT)):
            if lesson_arranged_count > 0 and create_time > last_week:
                arranged_count += 1
            if lesson_arranged_count == 0:
                un_arranged_count += 1
        return arranged_count, un_arranged_count

    def get_need_arrange(self):
        contract_ids = []
        arrange_count = 0

        for i in ex.bs("""select c.id from by_contract c
                            where c.create_time >= '{start}' and c.create_time < '{end}' 
                              and c.is_deleted=false and c.renew_from_contract_id is null
                              and c.store_id = {store_id}  and c.lesson_arraged_count = 0
                            """.format(store_id=self.store_id,
                                       start=date_param.start,
                                       end=date_param.end)):
            contract_ids.append(i[0])

        for i in ex.bs("""select DISTINCT c.id
                            from by_contract c
                                   left join by_contract_course_allocation bcca on bcca.contract_id = c.id
                                   left join by_student_lesson sl on sl.contract_course_allocation_id = bcca.id
                            where c.is_deleted = false
                              and c.store_id = {store_id} 
                              # and c.create_time >=date_add(curdate(),interval -7 day) and c.create_time < curdate()
                              and c.renew_from_contract_id is null
                              and sl.create_time >= '{start}'
                              and sl.create_time < '{end}'
                              and sl.is_deleted = false""".format(store_id=self.store_id,
                                                                  start=date_param.start,
                                                                  end=date_param.end)):
            contract_ids.append(i[0])
            arrange_count += 1
        return len(set(contract_ids)), arrange_count

    def get_ignore_mobile(self):
        data = []
        if self.store_name != "青浦青湖":
            return data
        for i in pe.load_book("doc/青浦结课老学员.xlsx").sheet_by_index(0):
            if i[1] == "手机号":
                continue
            data.append(i[1])
        return data

    def get_signed_time(self, contract_mobile, param):
        param.append(contract_mobile)
        array = [self.new_signed_mobile.get(i) for i in param if self.new_signed_mobile.get(i)]
        try:
            return min(array)
        except Exception:
            return datetime.datetime.now()


def signed_detail():
    buffer = BytesIO()

    all_signed_mobile = {i[0]: i[1]
                         for i in ex.qb("""select mobile_no,create_time from by_user where mobile_no is not null""")}
    new_signed_mobile = {i[0]: i[1] for i in ex.qb(
        """select mobile_no,create_time from by_user where create_time >='{start}' 
                and create_time < '{end}' and mobile_no is not null""".format(start=date_param.start,
                                                                              end=date_param.end))}

    end = (datetime.datetime.now() - datetime.timedelta(days=1)).date()
    start = (datetime.datetime.now() - datetime.timedelta(days=7)).date()

    wb = op.Workbook()
    del wb["Sheet"]
    for _id, _name in NEW_STORE_MAP.items():
        wb.create_sheet(_name)

        sheet = wb[_name]
        st = SignedTable(_id, _name, all_signed_mobile, new_signed_mobile)
        data = st.run()
        sheet["A1"] = "{name} {start} ~ {end} 报表".format(name=_name,
                                                         start=start.strftime("%Y-%m-%d"),
                                                         end=end.strftime("%Y-%m-%d"))
        sheet["A2"] = "学员端情况"
        sheet["A3"] = "老学员情况"
        sheet["B3"] = "应绑定"
        sheet["C3"] = "实绑定"
        sheet["D3"] = "本周老学员绑定数"
        sheet["E3"] = "绑定率"

        sheet["A6"] = "新签情况"
        sheet["B6"] = "新签合同数"
        sheet["C6"] = "学员端绑定数"
        sheet["D6"] = "绑定率"
        sheet["E6"] = "平均绑定天数"

        sheet["A8"] = "排课情况"
        sheet["A9"] = "新签数"
        sheet["B9"] = "本周新签已排课数"
        sheet["C9"] = "新签总计未排数"
        sheet["A12"] = "本周应排数"
        sheet["B12"] = "本周实际排课数"

        sheet["B4"] = data["学员端情况"]["老学员"]["应绑定"]
        sheet["C4"] = data["学员端情况"]["老学员"]["实绑定"]
        sheet["D4"] = data["学员端情况"]["老学员"]["本周老学员绑定数"]
        sheet["E4"] = "{}%".format(data["学员端情况"]["老学员"]["绑定率"] * 100)

        sheet["B7"] = data["学员端情况"]["新签情况"]["新签合同数"]
        sheet["C7"] = data["学员端情况"]["新签情况"]["学员端绑定数"]
        sheet["D7"] = "{}%".format(data["学员端情况"]["新签情况"]["绑定率"] * 100)
        sheet["E7"] = data["学员端情况"]["新签情况"]["平均绑定天数"]

        # sheet["A10"] = data["排课情况"]["新签数"]
        sheet["A10"] = data["学员端情况"]["新签情况"]["新签合同数"]
        sheet["B10"] = data["排课情况"]["本周新签已排课"]
        sheet["C10"] = data["排课情况"]["新签总计未排课"]

        sheet["A13"] = data["排课情况"]["本周应该排课"]
        sheet["B13"] = data["排课情况"]["本周实际排课"]

        font = Font(b=True)
        for line in [1, 2, 8]:
            sheet["A{line}".format(line=line)].font = font
            sheet.merge_cells("A{line}:E{line}".format(line=line))

        sheet.merge_cells("A3:A4")
        sheet.merge_cells("A6:A7")

        align = Alignment(horizontal='center', vertical='center')
        for col in sheet.columns:
            max_length = 0
            column = col[0].column  # Get the column name
            for cell in col:
                cell.alignment = align
                if cell.coordinate in sheet.merged_cells:  # not check merge_cells
                    continue
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 2.0
            sheet.column_dimensions[column].width = adjusted_width

    wb.save(buffer)
    wb.save("demo.xlsx")
    return buffer
